"""Business logic for the tracker tab: formatting, week aggregation, export."""
from __future__ import annotations

import csv
import datetime as dt
import io
from collections import deque
from typing import NamedTuple

from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from app.features.tracker import rating, schemas

# The user's dynamic ELO rating lives in rating.py (constants, anchor store,
# snapshots, handicap ladder, replay engine). Re-exported names keep existing
# call sites (router pieces, tests) working.
from app.features.tracker.rating import (  # noqa: F401
    compute_my_rating,
    get_my_anchor_date,
    get_my_points,
    handicap_bonus,
    set_my_points,
    snapshot_match_points,
)
from app.features.tracker.models import (
    Activity,
    Category,
    DayNote,
    Event,
    Match,
    PhysicalCheck,
    Player,
    Setting,
)
from app.features.training import service as training_service

# ---------------------------------------------------------------- physical items

# Fixed checklist for the Physical Training row. (key, English label)
PHYSICAL_ITEMS: list[tuple[str, str]] = [
    ("wall_sit", "Wall Sit"),
    ("sit_ups", "Sit-ups"),
    ("plank", "Plank"),
    ("squats", "Squats"),
    ("obliques", "Obliques"),
    ("stretching", "Stretching"),
]
PHYSICAL_ITEM_LABELS = dict(PHYSICAL_ITEMS)
# The Physical Training cell turns yellow once at least this share is ticked.
# Single source of truth lives in the Training Center (it applies the same rule
# to its own sessions) — re-exported here for the legacy-checklist path.
PHYSICAL_YELLOW_RATIO = training_service.PHYSICAL_YELLOW_RATIO


def physical_checks_by_date(checks: list[PhysicalCheck]) -> dict[str, list[str]]:
    """Group ticked item keys per ISO date, in the canonical item order."""
    order = {key: i for i, (key, _) in enumerate(PHYSICAL_ITEMS)}
    grouped: dict[str, list[str]] = {}
    for c in checks:
        grouped.setdefault(c.date.isoformat(), []).append(c.item_key)
    for iso in grouped:
        grouped[iso].sort(key=lambda k: order.get(k, 999))
    return grouped


def format_physical_cell(item_keys: list[str]) -> str:
    """Render ticked items as their labels, separated by a middot divider."""
    return " · ".join(PHYSICAL_ITEM_LABELS.get(k, k) for k in item_keys)


def physical_is_yellow(item_keys: list[str]) -> bool:
    if not PHYSICAL_ITEMS:
        return False
    return len(item_keys) / len(PHYSICAL_ITEMS) >= PHYSICAL_YELLOW_RATIO


# Max characters shown for a note in the (compact) grid cell.
_NOTE_SNIPPET_LEN = 22

# Coaching packages: a block of N sessions; ★ marks the first session of a block.
COACH_PACKAGE_SIZE = 10
PACKAGE_MARK = "★"

# ------------------------------------------------------------- racket time
# "Racket Time" = total time with the racket in hand per day: the coach +
# partner training minutes plus the match play estimated from set counts
# (a set averages ~5 minutes). Serve practice / physical work don't count.
RACKET_MINUTES_PER_SET = 5
RACKET_DURATION_KEYS = ("train_with_coach", "training_with_partner")


def racket_minutes_by_day(
    categories: list["Category"],
    activities: list["Activity"],
    matches: list["Match"],
) -> tuple[dict[str, int], dict[str, int]]:
    """Per ISO day: (training minutes from coach+partner, match minutes from
    sets × RACKET_MINUTES_PER_SET). Non-playing entries (Travel/Rest) don't
    count. Racket time for a day = the sum of the two maps' values."""
    racket_ids = {c.id for c in categories if c.key in RACKET_DURATION_KEYS}
    training: dict[str, int] = {}
    playing: dict[str, int] = {}
    for a in activities:
        if a.category_id in racket_ids and (a.duration_minutes or 0) > 0:
            iso = a.date.isoformat()
            training[iso] = training.get(iso, 0) + a.duration_minutes
    for m in matches:
        if m.is_nonplaying:
            continue
        sets = (m.my_sets or 0) + (m.opp_sets or 0)
        if sets <= 0:
            continue
        iso = m.date.isoformat()
        playing[iso] = playing.get(iso, 0) + sets * RACKET_MINUTES_PER_SET
    return training, playing


def note_snippet(text: str) -> str:
    """A compact one-line preview for the grid cell ('📝 ...'). Full text lives
    in the editor / export."""
    s = " ".join((text or "").split())
    if len(s) > _NOTE_SNIPPET_LEN:
        s = s[:_NOTE_SNIPPET_LEN].rstrip() + "…"
    return f"📝 {s}".rstrip()


# ---------------------------------------------------------------- formatting


def format_duration(minutes: int) -> str:
    """60 -> '1 hour', 45 -> '45 mins', 90 -> '1 hour 30 mins'."""
    if not minutes:
        return ""
    hours, mins = divmod(minutes, 60)
    parts: list[str] = []
    if hours:
        parts.append(f"{hours} hour")
    if mins:
        parts.append(f"{mins} mins")
    return " ".join(parts)


def _result_letter(my_sets: int, opp_sets: int) -> str:
    if my_sets > opp_sets:
        return "W"
    if my_sets < opp_sets:
        return "L"
    return "T"


# Fixed ordering of result groups within a cell.
_GROUP_ORDER = [
    (d, r)
    for d in ("singles", "doubles", "one_v_two", "two_v_one")
    for r in ("W", "L", "T")
]

# Cell/export prefix per discipline (singles stays bare, Excel-style).
_DISCIPLINE_PREFIX = {"doubles": "D: ", "one_v_two": "1v2: ", "two_v_one": "2v1: "}


def format_match_cell(matches: list[Match]) -> str:
    """Render a day's matches the way the Excel sheet shows them.

    e.g. 'W(3-0,3-1)' / 'D: L(2-3,1-3)', with event names and non-playing
    labels (Travel/Rest) on their own lines.
    """
    if not matches:
        return ""

    ordered = sorted(matches, key=lambda m: m.order_index)
    lines: list[str] = []

    # Non-playing markers (Travel / Rest).
    nonplaying: list[str] = []
    for m in ordered:
        if m.is_nonplaying and m.nonplaying_label and m.nonplaying_label not in nonplaying:
            nonplaying.append(m.nonplaying_label)

    # Event names (distinct, first-seen order).
    events: list[str] = []
    for m in ordered:
        if not m.is_nonplaying and m.event and m.event.name and m.event.name not in events:
            events.append(m.event.name)

    # Group scores by (discipline, result).
    groups: dict[tuple[str, str], list[str]] = {}
    for m in ordered:
        if m.is_nonplaying:
            continue
        key = (m.discipline, _result_letter(m.my_sets, m.opp_sets))
        groups.setdefault(key, []).append(f"{m.my_sets}-{m.opp_sets}")

    lines.extend(nonplaying)
    lines.extend(events)

    ordered_keys = _GROUP_ORDER + [k for k in groups if k not in _GROUP_ORDER]
    for key in ordered_keys:
        scores = groups.get(key)
        if not scores:
            continue
        discipline, result = key
        prefix = _DISCIPLINE_PREFIX.get(discipline, "")
        lines.append(f"{prefix}{result}({','.join(scores)})")

    return "\n".join(lines)


def match_to_out(m: Match, my_points: int) -> schemas.MatchOut:
    """`my_points` = my CURRENT dynamic rating — the *_level fields are
    derived from points (at-match-time snapshot first), not the retired
    hand-picked label."""
    return schemas.MatchOut(
        id=m.id,
        date=m.date,
        category_id=m.category_id,
        discipline=m.discipline,
        best_of=m.best_of,
        my_sets=m.my_sets,
        opp_sets=m.opp_sets,
        event_id=m.event_id,
        event_name=m.event.name if m.event else None,
        is_nonplaying=m.is_nonplaying,
        nonplaying_label=m.nonplaying_label,
        note=m.note,
        order_index=m.order_index,
        opponent_id=m.opponent_id,
        opponent_name=m.opponent.name if m.opponent else None,
        opponent_level=_level_of(m.opp_points_snap, m.opponent, my_points),
        opponent_plays_pips=bool(m.opponent.plays_pips) if m.opponent else False,
        opponent2_id=m.opponent2_id,
        opponent2_name=m.opponent2.name if m.opponent2 else None,
        opponent2_level=_level_of(m.opp2_points_snap, m.opponent2, my_points),
        opponent2_plays_pips=bool(m.opponent2.plays_pips) if m.opponent2 else False,
        partner_id=m.partner_id,
        partner_name=m.partner.name if m.partner else None,
        partner_level=_level_of(m.partner_points_snap, m.partner, my_points),
        handicap=m.handicap or 0,
        handicap_pattern=m.handicap_pattern,
    )


# ---------------------------------------------------------------- players




def player_to_out(p: Player) -> schemas.PlayerOut:
    return schemas.PlayerOut(
        id=p.id,
        name=p.name,
        level=p.level,
        note=p.note,
        plays_pips=bool(p.plays_pips),
        points=p.points,
    )


def list_players(db: Session, q: str = "") -> list[schemas.PlayerOut]:
    query = db.query(Player)
    if q:
        query = query.filter(Player.name.ilike(f"%{q}%"))
    rows = query.order_by(Player.name).limit(50).all()
    return [player_to_out(p) for p in rows]


# Rank bands (H=0, then G..A per 200 points) — mirrors frontend shared/rank.ts.
_RANK_TOPS = (1000, 1200, 1400, 1600, 1800, 2000)


def _rank_band(points: int) -> int:
    if points < 800:
        return 0  # H
    for band, top in enumerate(_RANK_TOPS, start=1):
        if points <= top:
            return band  # G..B
    return 7  # A


def level_from_points(points: int | None, my_points: int) -> str:
    """Relative level derived from POINTS (retired the hand-picked label,
    2026-07-27): same rank band as my current dynamic rating = equal;
    no points yet = its own "unrated" bucket."""
    if points is None:
        return "unrated"
    theirs, mine = _rank_band(points), _rank_band(my_points)
    return "above" if theirs > mine else "below" if theirs < mine else "equal"


def _level_of(snap: int | None, player: Player | None, my_points: int) -> str | None:
    """Level of one match slot — at-match-time snapshot first, like the ELO."""
    if player is None:
        return None
    return level_from_points(snap if snap is not None else player.points, my_points)


def create_or_get_player(db: Session, payload: schemas.PlayerIn) -> schemas.PlayerOut:
    """Get-or-create by name. If the player exists, keep it.

    `level` is FROZEN legacy (2026-07-27): no longer written from payloads —
    analytics derive the relative level from points instead. New rows get the
    column default; the field is still accepted so old clients don't break."""
    name = (payload.name or "").strip()
    existing = db.query(Player).filter(Player.name == name).first()
    if existing is None:
        existing = Player(
            name=name,
            note=payload.note,
            plays_pips=payload.plays_pips,
            points=payload.points,
        )
        db.add(existing)
        db.commit()
        db.refresh(existing)
    return player_to_out(existing)


def update_player(
    db: Session, player_id: int, payload: schemas.PlayerIn
) -> schemas.PlayerOut | None:
    p = db.get(Player, player_id)
    if p is None:
        return None
    # Renaming is safe by design: matches reference players by id, so every
    # historical display (grid, h2h, coach bundle) picks up the new name on
    # the next read. Renaming INTO an existing player's name is rejected —
    # two identical rows would be indistinguishable in the picker (a real
    # "same person twice" situation needs a merge feature, not a rename).
    new_name = (payload.name or "").strip()
    if new_name and new_name.lower() != (p.name or "").lower():
        clash = (
            db.query(Player)
            .filter(func.lower(Player.name) == new_name.lower(), Player.id != player_id)
            .first()
        )
        if clash is not None:
            raise ValueError(f'A player named "{clash.name}" already exists.')
    p.name = new_name or p.name
    # payload.level is ignored — the column is frozen legacy (see above).
    p.note = payload.note
    p.plays_pips = payload.plays_pips
    # None = caller doesn't manage points (e.g. the picker's pips toggle) —
    # never wipe a rating the user set in the Database tab.
    if payload.points is not None:
        p.points = payload.points
    db.commit()
    db.refresh(p)
    return player_to_out(p)


def list_players_db(db: Session) -> schemas.PlayersDbResponse:
    """Every player + how often they appear in matches (the Database tab).

    Rated players first (highest points on top — a ranking table), unrated
    last alphabetically so they're easy to work through."""
    players = db.query(Player).all()
    # Two separate tallies: facing me (either opponent slot) vs on my side.
    vs_counts: dict[int, int] = {}
    with_counts: dict[int, int] = {}
    for col, counts in (
        (Match.opponent_id, vs_counts),
        (Match.opponent2_id, vs_counts),
        (Match.partner_id, with_counts),
    ):
        rows = (
            db.query(col, func.count(Match.id))
            # Playing matches only — the same rule as the per-player
            # drill-down, so a count badge always equals its modal rows.
            .filter(col.isnot(None), Match.is_nonplaying == False)  # noqa: E712
            .group_by(col)
            .all()
        )
        for pid, n in rows:
            counts[pid] = counts.get(pid, 0) + n
    ordered = sorted(
        players,
        key=lambda p: (
            p.points is None,  # rated first
            -(p.points or 0),
            p.name.lower(),
        ),
    )
    return schemas.PlayersDbResponse(
        players=[
            schemas.PlayerDbRow(
                **player_to_out(p).model_dump(),
                matches_vs=vs_counts.get(p.id, 0),
                matches_with=with_counts.get(p.id, 0),
            )
            for p in ordered
        ]
    )


# ---------------------------------------------------------------- events


def get_or_create_event(db: Session, name: str | None) -> Event | None:
    name = (name or "").strip()
    if not name:
        return None
    event = db.query(Event).filter(Event.name == name).first()
    if event is None:
        event = Event(name=name)
        db.add(event)
        db.flush()
    return event


# ---------------------------------------------------------------- overall color


def earliest_data_date(db: Session) -> dt.date | None:
    """The first day that has any tracked data (for the red 'rest-day' bound)."""
    tc_lo, _ = training_service.done_date_bounds(db)
    candidates = [
        db.query(func.min(Activity.date)).scalar(),
        db.query(func.min(Match.date)).scalar(),
        db.query(func.min(PhysicalCheck.date)).scalar(),
        tc_lo,
    ]
    dates = [d for d in candidates if d is not None]
    return min(dates) if dates else None


def latest_data_date(db: Session) -> dt.date | None:
    """The most recent day that has any tracked data (for opening the grid there)."""
    _, tc_hi = training_service.done_date_bounds(db)
    candidates = [
        db.query(func.max(Activity.date)).scalar(),
        db.query(func.max(Match.date)).scalar(),
        db.query(func.max(PhysicalCheck.date)).scalar(),
        tc_hi,
    ]
    dates = [d for d in candidates if d is not None]
    return max(dates) if dates else None


def compute_overall_colors(
    categories: list[Category],
    activities: list[Activity],
    matches: list[Match],
    physical_dates: set[str] | None = None,
    all_days: list[dt.date] | None = None,
    today: dt.date | None = None,
    earliest: dt.date | None = None,
) -> dict[str, str]:
    """Auto-generate the 'Overall' color per day (no manual rating).

    - green: any of the green-group rows (Train with Coach / Backhand with
      Partner / Serve) has duration data that day.
    - yellow: otherwise, any of the remaining rows (Physical Training, Practice
      Match, Official Match) has data that day.
    - red: a past day (>= the first tracked day, < today) with no data at all.
    - (absent): no data, and the day is today, in the future, or before tracking
      began.

    Returns {iso_date: 'green' | 'yellow' | 'red'}.
    """
    green_ids = {c.id for c in categories if c.color_group == "green"}

    green_days: set[str] = set()
    other_days: set[str] = set(physical_dates or set())

    for a in activities:
        if (a.duration_minutes or 0) <= 0:
            continue
        iso = a.date.isoformat()
        if a.category_id in green_ids:
            green_days.add(iso)
        else:
            other_days.add(iso)

    # Any match entry (including Travel/Rest) counts as activity for the day.
    for m in matches:
        other_days.add(m.date.isoformat())

    colors: dict[str, str] = {}
    for iso in green_days | other_days:
        colors[iso] = "green" if iso in green_days else "yellow"

    # Empty past days within the tracked range -> red ("didn't train").
    if all_days and today:
        for d in all_days:
            iso = d.isoformat()
            if iso in colors:
                continue
            if earliest is not None and earliest <= d < today:
                colors[iso] = "red"
    return colors


# ---------------------------------------------------------------- range loading


class RangeData(NamedTuple):
    categories: list[Category]
    activities: list[Activity]
    matches: list[Match]  # ordered by order_index
    checks_by_date: dict[str, list[str]]  # iso date -> ticked item keys (legacy)
    notes_by_date: dict[str, str]  # iso date -> day-note text
    # Training Center sessions completed in-range, keyed by done date. From the
    # cutover forward these are the source of the physical-training signal.
    tc_physical: dict[str, dict]


def _load_range(
    db: Session,
    date_from: dt.date,
    date_to: dt.date,
    with_match_relations: bool = True,
) -> RangeData:
    """Load the four range-filtered tables shared by the week / stats /
    breakdown / export builders, so each one stops re-issuing the same queries.

    ``with_match_relations`` eager-loads each match's event/opponent/partner via
    selectinload (avoids N+1 where they're rendered); callers that only count
    matches (stats / breakdown) pass False to skip the extra queries.
    """
    categories = db.query(Category).order_by(Category.sort_order).all()
    activities = (
        db.query(Activity)
        .filter(Activity.date >= date_from, Activity.date <= date_to)
        .all()
    )
    match_q = db.query(Match).filter(Match.date >= date_from, Match.date <= date_to)
    if with_match_relations:
        match_q = match_q.options(
            selectinload(Match.event),
            selectinload(Match.opponent),
            selectinload(Match.opponent2),
            selectinload(Match.partner),
        )
    matches = match_q.order_by(Match.order_index).all()
    checks = (
        db.query(PhysicalCheck)
        .filter(PhysicalCheck.date >= date_from, PhysicalCheck.date <= date_to)
        .all()
    )
    day_notes = (
        db.query(DayNote)
        .filter(DayNote.date >= date_from, DayNote.date <= date_to)
        .all()
    )
    return RangeData(
        categories=categories,
        activities=activities,
        matches=matches,
        checks_by_date=physical_checks_by_date(checks),
        notes_by_date={n.date.isoformat(): n.text for n in day_notes},
        tc_physical=training_service.physical_day_map(db, date_from, date_to),
    )


def _physical_dates(rng: RangeData) -> set[str]:
    """All dates that count as a physical-training day: legacy checklist ticks
    (before the cutover) unioned with completed Training Center sessions (from
    the cutover forward). The two never overlap, so this can't double-count."""
    return set(rng.checks_by_date.keys()) | set(rng.tc_physical.keys())


# ---------------------------------------------------------------- grid cells


def _grid_cells(
    db: Session, rng: RangeData, days: list[dt.date], *, for_export: bool
) -> tuple[dict[tuple[int, str], str], dict[tuple[int, str], str], dict[str, str]]:
    """The ONE cell renderer behind both the on-screen grid (build_week) and
    the export (_build_grid). They used to be two ~70-line near-copies that
    drifted (the export-parity bug came from exactly that), so every row type
    renders here exactly once.

    Returns (text, colors, overall_color_by_date); text/colors are keyed by
    (category_id, iso date). ``for_export``: day notes render in full (the
    screen shows the compact 📝 snippet; full text travels in day_notes) and
    the Training-Center mirror uses words instead of the 💪 emoji.
    """
    categories = rng.categories
    cat_by_key = {c.key: c for c in categories}
    duration_ids = {c.id for c in categories if c.type == "duration"}
    activities, matches = rng.activities, rng.matches

    text: dict[tuple[int, str], str] = {}
    colors: dict[tuple[int, str], str] = {}

    # Duration cells (sum minutes per category/day) + note/★ suffixes.
    mins: dict[tuple[int, str], int] = {}
    cell_notes: dict[tuple[int, str], str] = {}  # per-cell activity note
    starts: set[tuple[int, str]] = set()  # cells whose day starts a package
    for a in activities:
        if a.category_id not in duration_ids:
            continue
        key = (a.category_id, a.date.isoformat())
        mins[key] = mins.get(key, 0) + (a.duration_minutes or 0)
        if a.note:
            cell_notes[key] = a.note
        if a.is_package_start:
            starts.add(key)
    for key, m in mins.items():
        cell = format_duration(m)
        if key in cell_notes:
            cell = f"{cell} ({cell_notes[key]})".strip()
        if key in starts:  # first session of a new 10-session coaching package
            cell = f"{cell} {PACKAGE_MARK}".strip()
        text[key] = cell

    # Match cells.
    by_cell: dict[tuple[int, str], list[Match]] = {}
    for mt in matches:
        by_cell.setdefault((mt.category_id, mt.date.isoformat()), []).append(mt)
    for key, ms in by_cell.items():
        text[key] = format_match_cell(ms)

    # Physical Training cells. Before the cutover: the legacy checklist (ticked
    # labels, yellow at >=70%). From the cutover forward: a read-only mirror of
    # the Training Center session done that day.
    physical = cat_by_key.get("physical_training")
    if physical is not None:
        for iso, keys in rng.checks_by_date.items():
            text[(physical.id, iso)] = format_physical_cell(keys)
            if physical_is_yellow(keys):
                colors[(physical.id, iso)] = "yellow"
        tc_prefix = "Training Center" if for_export else "💪"
        for iso, info in rng.tc_physical.items():
            text[(physical.id, iso)] = (
                f"{tc_prefix} {info['done']}/{info['total']} · {info['focus_vi']}"
            )
            if info["is_yellow"]:
                colors[(physical.id, iso)] = "yellow"

    # Racket Time cells: auto-computed (coach + partner + 5 min per match set).
    racket = cat_by_key.get("racket_time")
    if racket is not None:
        r_training, r_playing = racket_minutes_by_day(categories, activities, matches)
        for iso in set(r_training) | set(r_playing):
            total = r_training.get(iso, 0) + r_playing.get(iso, 0)
            text[(racket.id, iso)] = format_duration(total)

    # Notes cells.
    notes_cat = cat_by_key.get("notes")
    if notes_cat is not None:
        for iso, note_text in rng.notes_by_date.items():
            text[(notes_cat.id, iso)] = (
                note_text if for_export else note_snippet(note_text)
            )

    # Overall: auto-generated colors from the day's data (not a manual rating).
    overall_colors = compute_overall_colors(
        categories,
        activities,
        matches,
        _physical_dates(rng),
        all_days=days,
        today=dt.date.today(),
        earliest=earliest_data_date(db),
    )
    return text, colors, overall_colors


# ---------------------------------------------------------------- week


def _annotate_elo(db: Session, matches: list[Match]) -> list[schemas.MatchOut]:
    """One replay pass → MatchOut per match, tagged with its ±Δ (counted) or
    the reason it doesn't move the rating (the GUI shows actionable "không
    tính"). The same pass yields my current rating for the derived *_level
    fields."""
    my_final, elo_steps = rating.replay(db)
    elo_deltas = {s.match_id: s.delta for s in elo_steps}
    anchor_date = rating.get_my_anchor_date(db)
    outs = [match_to_out(m, round(my_final)) for m in matches]
    for out, m in zip(outs, matches):
        if m.id in elo_deltas:
            out.elo_delta = round(elo_deltas[m.id], 1)
            out.elo_status = rating.STATUS_COUNTED
        else:
            out.elo_status = rating.skip_reason(m, anchor_date)
    return outs


def build_week(
    db: Session, start: dt.date, end: dt.date | None = None
) -> schemas.WeekResponse:
    # The grid can span an arbitrary range (a single day, a week, a whole
    # month, …). Defaults to a 7-day Mon–Sun week when no end is given.
    if end is None:
        end = start + dt.timedelta(days=6)
    if end < start:
        end = start
    span = (end - start).days + 1
    days = [start + dt.timedelta(days=i) for i in range(span)]

    rng = _load_range(db, start, end)
    categories = rng.categories
    cat_by_key = {c.key: c for c in categories}
    activities = rng.activities
    matches = rng.matches
    checks_by_date = rng.checks_by_date
    notes_by_date = rng.notes_by_date

    text, colors, overall_colors = _grid_cells(db, rng, days, for_export=False)
    cells: dict[str, schemas.CellData] = {
        f"{cid}|{iso}": schemas.CellData(
            display=display, color=colors.get((cid, iso))
        )
        for (cid, iso), display in text.items()
    }
    overall = cat_by_key.get("overall")
    if overall is not None:
        for iso, color in overall_colors.items():
            cells[f"{overall.id}|{iso}"] = schemas.CellData(display="", color=color)

    match_outs = _annotate_elo(db, matches)

    return schemas.WeekResponse(
        start=start,
        days=days,
        categories=[schemas.CategoryOut.model_validate(c) for c in categories],
        activities=[schemas.ActivityOut.model_validate(a) for a in activities],
        matches=match_outs,
        cells=cells,
        physical_checks=checks_by_date,
        day_notes=notes_by_date,
        physical_cutover=training_service.get_cutover(db),
    )


# ---------------------------------------------------------------- coach packages


def _coach_sessions(db: Session) -> list[Activity]:
    """Every Train-with-Coach session with real duration, in date order —
    the single query behind all package computations."""
    coach = db.query(Category).filter(Category.key == "train_with_coach").first()
    if coach is None:
        return []
    return (
        db.query(Activity)
        .filter(Activity.category_id == coach.id, Activity.duration_minutes > 0)
        .order_by(Activity.date)
        .all()
    )


def compute_coach_packages(db: Session) -> schemas.CoachPackagesResponse:
    """Group Train-with-Coach sessions into packages of COACH_PACKAGE_SIZE.

    A package opens on each session flagged is_package_start; the earliest
    session implicitly opens package #1 (covers data older than any marker).
    """
    sessions = _coach_sessions(db)
    size = COACH_PACKAGE_SIZE
    packages: list[schemas.CoachPackage] = []
    for i, a in enumerate(sessions):
        opens = a.is_package_start or i == 0
        if opens:
            packages.append(
                schemas.CoachPackage(
                    number=len(packages) + 1,
                    start_date=a.date,
                    end_date=a.date,
                    used=1,
                    size=size,
                    remaining=max(0, size - 1),
                    over=max(0, 1 - size),
                    is_current=False,
                    status="ok",
                )
            )
        else:
            p = packages[-1]
            p.used += 1
            p.end_date = a.date
            p.remaining = max(0, size - p.used)
            p.over = max(0, p.used - size)

    # Status for every package (history included), not just the current one —
    # a past block that ran over should read "over" in the history list too.
    for p in packages:
        p.is_current = False
        if p.over > 0:
            p.status = "over"
        elif p.remaining == 0:
            p.status = "done"
        elif p.remaining <= 2:
            p.status = "low"
        else:
            p.status = "ok"
    if packages:
        packages[-1].is_current = True

    return schemas.CoachPackagesResponse(size=size, packages=packages)


def start_next_coach_package(db: Session) -> schemas.CoachPackagesResponse:
    """One-click card action: when the current block ran over its size, flag
    its (size+1)-th session as the next package's first session.

    Equivalent to opening that day's coach cell and ticking the ★ box — this
    just finds the right day automatically (always session 11, so sessions
    12+ stay in the NEW package, never inflate the old one)."""
    sessions = _coach_sessions(db)
    # Current block = everything from the last flagged start (or session #1).
    start_idx = 0
    for i, a in enumerate(sessions):
        if a.is_package_start:
            start_idx = i
    block = sessions[start_idx:]
    if len(block) <= COACH_PACKAGE_SIZE:
        raise ValueError(
            f"The current package only has {len(block)}/{COACH_PACKAGE_SIZE} "
            "sessions — no session has run over the package yet."
        )
    block[COACH_PACKAGE_SIZE].is_package_start = True
    db.commit()
    return compute_coach_packages(db)


def coach_package_start_allowed(db: Session, date: dt.date) -> bool:
    """Whether `date` may be marked as the start of a new coaching package.

    Allowed only when the day is the first session of its package (position 1 —
    so an existing start can be un-marked) or the 11th-or-later session of the
    current block (i.e. the previous 10 sessions are used up). Sessions 2..10
    of a block are NOT allowed. Works for a date that has no session yet
    (e.g. logging the 11th session for the first time).
    """
    sessions = _coach_sessions(db)
    if not sessions:
        return True  # the very first session can always open package #1

    # Package start dates = flagged sessions, plus an implicit first session.
    starts = sorted({s.date for s in sessions if s.is_package_start} | {sessions[0].date})

    applicable = [s for s in starts if s <= date]
    if not applicable:
        return True  # before the first start -> would become the new earliest start
    pkg_start = max(applicable)
    later = [s for s in starts if s > pkg_start]
    pkg_end = min(later) if later else None  # exclusive upper bound

    # Position of `date` within its block = (sessions in block before `date`) + 1.
    before = sum(
        1
        for s in sessions
        if s.date >= pkg_start
        and (pkg_end is None or s.date < pkg_end)
        and s.date < date
    )
    pos = before + 1
    return pos == 1 or pos >= COACH_PACKAGE_SIZE + 1


# ---------------------------------------------------------------- stats

_WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _blank_match_stats() -> dict:
    return {"total": 0, "wins": 0, "losses": 0, "ties": 0, "sets_won": 0, "sets_lost": 0}


def win_rate(wins: int, losses: int) -> float | None:
    """Win rate over *decided* matches (ties excluded). None if none decided."""
    decided = wins + losses
    return (wins / decided) if decided else None


def normalize_handicap_pattern(raw: str | None) -> str | None:
    """"202" / "2-0-2" → "2-0-2". Uniform sequences ("222") and empty input
    collapse to None — the plain signed `handicap` int already carries a
    uniform per-set value, so only genuinely mixed ratios store a pattern."""
    digits = [c for c in (raw or "") if c.isdigit()]
    if not digits or len(set(digits)) == 1:
        return None
    return "-".join(digits)


def last_handicap_vs(db: Session, player_id: int) -> Match | None:
    """The most recent singles playing match against `player_id` — its
    handicap is the best guess for the next match's ratio (the GUI pre-fills
    it when the opponent is picked; the user can still change it)."""
    return (
        db.query(Match)
        .filter(
            Match.opponent_id == player_id,
            Match.discipline == "singles",
            Match.is_nonplaying == False,  # noqa: E712
        )
        .order_by(Match.date.desc(), Match.order_index.desc(), Match.id.desc())
        .first()
    )


def _result_of(m: Match) -> str:
    return _result_letter(m.my_sets, m.opp_sets)


def _tally(s: dict, m: Match) -> None:
    """Fold one playing match into a _blank_match_stats() accumulator."""
    s["total"] += 1
    s["sets_won"] += m.my_sets
    s["sets_lost"] += m.opp_sets
    r = _result_of(m)
    if r == "W":
        s["wins"] += 1
    elif r == "L":
        s["losses"] += 1
    else:
        s["ties"] += 1


def _finalize_match_stats(s: dict) -> schemas.MatchStats:
    return schemas.MatchStats(**s, win_rate=win_rate(s["wins"], s["losses"]))


def _in_stats_bucket(m: Match, bucket: str) -> bool:
    """Whether a match belongs to one of build_stats' summary buckets.

    The SAME predicate backs the stat cards (build_stats) and the drill-down
    list (list_stats_matches), so a card's numbers and its match list can
    never disagree."""
    if m.is_nonplaying:
        return False
    if bucket == "overall":
        return True
    if bucket == "vs_pips":
        # Either listed opponent plays pimpled rubber (covers doubles).
        return bool(
            (m.opponent and m.opponent.plays_pips)
            or (m.opponent2 and m.opponent2.plays_pips)
        )
    if bucket == "singles":
        return m.discipline not in ("doubles", "one_v_two", "two_v_one")
    return m.discipline == bucket


def list_stats_matches(
    db: Session, date_from: dt.date, date_to: dt.date, bucket: str
) -> list[schemas.MatchOut]:
    """Drill-down behind one stat card: the matches making up that bucket's
    numbers in the range, newest first, ELO-annotated like the week view."""
    rng = _load_range(db, date_from, date_to, with_match_relations=True)
    matches = [m for m in rng.matches if _in_stats_bucket(m, bucket)]
    matches.sort(key=lambda m: (m.date, m.order_index, m.id), reverse=True)
    return _annotate_elo(db, matches)


# SQL ordering for "newest first" — order_index then id break same-day ties.
_NEWEST_FIRST = (Match.date.desc(), Match.order_index.desc(), Match.id.desc())


def _playing_matches(db: Session, with_relations: bool = True):
    """Base query for playing matches (nonplaying rows excluded), optionally
    eager-loading the full line-up + event so names resolve without N+1."""
    q = db.query(Match).filter(Match.is_nonplaying == False)  # noqa: E712
    if with_relations:
        q = q.options(
            selectinload(Match.event),
            selectinload(Match.opponent),
            selectinload(Match.opponent2),
            selectinload(Match.partner),
        )
    return q


def list_player_matches(db: Session, player_id: int) -> list[schemas.MatchOut]:
    """Every playing match involving `player_id` in ANY slot (opponent,
    second opponent, or my partner) — the Database tab's per-player
    drill-down. All-time, newest first, ELO-annotated like the week view."""
    matches = (
        _playing_matches(db)
        .filter(
            or_(
                Match.opponent_id == player_id,
                Match.opponent2_id == player_id,
                Match.partner_id == player_id,
            )
        )
        .order_by(*_NEWEST_FIRST)
        .all()
    )
    return _annotate_elo(db, matches)


def build_stats(db: Session, date_from: dt.date, date_to: dt.date) -> schemas.StatsResponse:
    dates = _date_range(date_from, date_to)
    iso_dates = [d.isoformat() for d in dates]

    # with_match_relations=True so opponent.plays_pips is available for the
    # "vs đối thủ đánh gai" split (the opponent records are eager-loaded).
    rng = _load_range(db, date_from, date_to, with_match_relations=True)
    duration_cats = [c for c in rng.categories if c.type == "duration"]
    duration_ids = {c.id for c in duration_cats}
    activities = rng.activities
    matches = rng.matches

    # Per-day aggregates.
    minutes_per_day: dict[str, int] = {iso: 0 for iso in iso_dates}
    physical_per_day: dict[str, int] = {iso: 0 for iso in iso_dates}
    matches_per_day: dict[str, int] = {iso: 0 for iso in iso_dates}

    minutes_by_cat: dict[int, int] = {cid: 0 for cid in duration_ids}
    for a in activities:
        if a.category_id not in duration_ids or (a.duration_minutes or 0) <= 0:
            continue
        iso = a.date.isoformat()
        minutes_per_day[iso] = minutes_per_day.get(iso, 0) + a.duration_minutes
        minutes_by_cat[a.category_id] += a.duration_minutes

    for iso, keys in rng.checks_by_date.items():
        physical_per_day[iso] = physical_per_day.get(iso, 0) + len(keys)
    # Training Center sessions (from the cutover forward) also count as physical
    # days — mark the date present so days_physical / "trained" pick it up.
    for iso in rng.tc_physical:
        physical_per_day[iso] = physical_per_day.get(iso, 0) + 1

    # Match stats (playing matches only), split by discipline.
    overall = _blank_match_stats()
    singles = _blank_match_stats()
    doubles = _blank_match_stats()
    one_v_two = _blank_match_stats()  # me alone vs two opponents
    two_v_one = _blank_match_stats()  # me + partner vs one opponent
    vs_pips = _blank_match_stats()  # matches against a pimpled-rubber opponent

    by_discipline = {
        "doubles": doubles,
        "one_v_two": one_v_two,
        "two_v_one": two_v_one,
    }
    for m in matches:
        if m.is_nonplaying:
            continue
        iso = m.date.isoformat()
        matches_per_day[iso] = matches_per_day.get(iso, 0) + 1

        bucket = by_discipline.get(m.discipline, singles)
        _tally(overall, m)
        _tally(bucket, m)
        if _in_stats_bucket(m, "vs_pips"):
            _tally(vs_pips, m)

    # Day-level counts.
    days_trained = sum(
        1
        for iso in iso_dates
        if minutes_per_day[iso] > 0
        or physical_per_day[iso] > 0
        or matches_per_day[iso] > 0
    )
    days_physical = sum(1 for iso in iso_dates if physical_per_day[iso] > 0)

    # In category display order (duration_cats already sorted by sort_order).
    minutes_by_category = [
        schemas.CategoryMinutes(
            key=c.key, label=c.label, minutes=minutes_by_cat.get(c.id, 0)
        )
        for c in duration_cats
    ]

    # Racket time over the range (training with racket + match play from sets).
    r_training, r_playing = racket_minutes_by_day(rng.categories, activities, matches)
    racket_training = sum(r_training.values())
    racket_matches = sum(r_playing.values())

    return schemas.StatsResponse(
        date_from=date_from,
        date_to=date_to,
        num_days=len(dates),
        days_trained=days_trained,
        days_physical=days_physical,
        minutes_total=sum(minutes_per_day.values()),
        minutes_by_category=minutes_by_category,
        racket_minutes_total=racket_training + racket_matches,
        racket_minutes_training=racket_training,
        racket_minutes_matches=racket_matches,
        overall=_finalize_match_stats(overall),
        singles=_finalize_match_stats(singles),
        doubles=_finalize_match_stats(doubles),
        one_v_two=_finalize_match_stats(one_v_two),
        two_v_one=_finalize_match_stats(two_v_one),
        vs_pips=_finalize_match_stats(vs_pips),
    )


# ---------------------------------------------------------------- breakdown


def _first_of_month(d: dt.date) -> dt.date:
    return d.replace(day=1)


def _next_month(d: dt.date) -> dt.date:
    return dt.date(d.year + (d.month // 12), (d.month % 12) + 1, 1)


def _last_of_month(d: dt.date) -> dt.date:
    return _next_month(d) - dt.timedelta(days=1)


def _bucket_ranges(date_from: dt.date, date_to: dt.date, unit: str):
    """Yield (key, label, b_from, b_to) sub-ranges that tile [from, to]."""
    if unit == "day":
        for d in _date_range(date_from, date_to):
            yield (d.isoformat(), f"{_WEEKDAYS[d.weekday()]} {d.day}", d, d)
        return
    if unit == "week":
        cur = date_from - dt.timedelta(days=date_from.weekday())  # Monday
        n = 0
        while cur <= date_to:
            n += 1
            w_end = cur + dt.timedelta(days=6)
            b_from = max(cur, date_from)
            b_to = min(w_end, date_to)
            yield (cur.isoformat(), f"Week {n}", b_from, b_to)
            cur = cur + dt.timedelta(days=7)
        return
    # default: month
    cur = _first_of_month(date_from)
    while cur <= date_to:
        m_end = _last_of_month(cur)
        b_from = max(cur, date_from)
        b_to = min(m_end, date_to)
        yield (cur.strftime("%Y-%m"), cur.strftime("%b"), b_from, b_to)
        cur = _next_month(cur)


def build_breakdown(
    db: Session, date_from: dt.date, date_to: dt.date, unit: str
) -> schemas.BreakdownResponse:
    """Per-sub-period metrics for the comparison bar chart."""
    rng = _load_range(db, date_from, date_to, with_match_relations=False)
    activities = rng.activities
    matches = rng.matches
    duration_ids = {c.id for c in rng.categories if c.type == "duration"}

    # Per-day aggregates.
    minutes: dict[str, int] = {}
    physical: set[str] = _physical_dates(rng)
    mcount: dict[str, int] = {}
    wins: dict[str, int] = {}
    losses: dict[str, int] = {}

    for a in activities:
        if a.category_id in duration_ids and (a.duration_minutes or 0) > 0:
            iso = a.date.isoformat()
            minutes[iso] = minutes.get(iso, 0) + a.duration_minutes
    for m in matches:
        if m.is_nonplaying:
            continue
        iso = m.date.isoformat()
        mcount[iso] = mcount.get(iso, 0) + 1
        r = _result_of(m)
        if r == "W":
            wins[iso] = wins.get(iso, 0) + 1
        elif r == "L":
            losses[iso] = losses.get(iso, 0) + 1

    buckets: list[schemas.BreakdownBucket] = []
    for key, label, b_from, b_to in _bucket_ranges(date_from, date_to, unit):
        b_min = b_w = b_l = b_m = b_trained = b_phys = 0
        for d in _date_range(b_from, b_to):
            iso = d.isoformat()
            mins = minutes.get(iso, 0)
            mc = mcount.get(iso, 0)
            ph = iso in physical
            b_min += mins
            b_m += mc
            b_w += wins.get(iso, 0)
            b_l += losses.get(iso, 0)
            if ph:
                b_phys += 1
            if mins > 0 or ph or mc > 0:
                b_trained += 1
        buckets.append(
            schemas.BreakdownBucket(
                key=key,
                label=label,
                date_from=b_from,
                date_to=b_to,
                minutes=b_min,
                days_trained=b_trained,
                days_physical=b_phys,
                matches=b_m,
                wins=b_w,
                losses=b_l,
                win_rate=win_rate(b_w, b_l),
            )
        )

    return schemas.BreakdownResponse(unit=unit, buckets=buckets)


# ------------------------------------------- match stats (Profile tab, middle)

# Relative levels are DERIVED from points (vs my current dynamic rating);
# "unrated" = opponent has no points yet. Hand-picked labels are retired.
_LEVEL_ORDER = ["below", "equal", "above", "unrated"]
_CATEGORY_KEY = {
    "practice": "practice_match",
    "official": "official_match",
    "tournament": "tournament_match",
}
# Match tracking with opponents began June 2026; ignore anything before.
MATCH_STATS_FLOOR = dt.date(2026, 6, 1)


def _query_named_matches(
    db: Session,
    date_from: dt.date,
    date_to: dt.date,
    discipline: str,
    category: str,
    with_relations: bool = True,
) -> list[Match]:
    """Playing matches with a named opponent in the range, in play order
    (order_index breaks same-day ties so "last_result" is truly the last).
    `with_relations=False` skips the four eager loads for callers that only
    read scores (the rolling-form seed)."""
    q = _playing_matches(db, with_relations).filter(
        Match.date >= date_from,
        Match.date <= date_to,
        Match.opponent_id.isnot(None),
    )
    if discipline in ("singles", "doubles", "one_v_two", "two_v_one"):
        q = q.filter(Match.discipline == discipline)
    if category in _CATEGORY_KEY:
        cat = db.query(Category).filter(Category.key == _CATEGORY_KEY[category]).first()
        q = q.filter(Match.category_id == (cat.id if cat else -1))
    return q.order_by(Match.date, Match.order_index).all()


def _record_tail(rec: dict) -> dict:
    """The stat fields OpponentRecord and DoublesRecord share."""
    return {
        "played": rec["total"],
        "wins": rec["wins"],
        "losses": rec["losses"],
        "ties": rec["ties"],
        "sets_won": rec["sets_won"],
        "sets_lost": rec["sets_lost"],
        "win_rate": win_rate(rec["wins"], rec["losses"]),
        "last_date": rec["last_date"],
        "last_result": rec["last_result"],
        "matches": list(reversed(rec.get("matches", []))),  # most recent first
    }


class _H2HAcc(NamedTuple):
    overall: dict
    singles_h2h: dict[int, dict]
    doubles_h2h: dict[str, dict]
    opp_brief: dict[int, dict]


def _h2h_accumulate(matches: list[Match], my_now: int) -> _H2HAcc:
    """One pass over the matches → overall tally + head-to-head records
    (singles per opponent; team-style per discipline+partner+pair)."""
    overall = _blank_match_stats()
    singles_h2h: dict[int, dict] = {}  # keyed by opponent_id
    # Team-style matchups (doubles / 1v2 / 2v1), keyed by
    # discipline + partner + opponent-pair. Slots a format doesn't use stay None.
    doubles_h2h: dict[str, dict] = {}
    opp_brief: dict[int, dict] = {}  # every opponent seen -> {name, level, played}

    for m in matches:
        _tally(overall, m)

        # Dropdown list: count every opponent appearance (opp1 + opp2).
        for opp in (m.opponent, m.opponent2):
            if opp is not None:
                b = opp_brief.get(opp.id)
                if b is None:
                    b = opp_brief[opp.id] = {
                        "id": opp.id,
                        "name": opp.name,
                        "level": level_from_points(opp.points, my_now),
                        "played": 0,
                    }
                b["played"] += 1

        if m.discipline != "singles":
            # A team-style matchup = (my partner, if any) vs (their unordered
            # opponent pair) — covers doubles, 1v2 (no partner) and 2v1 (one
            # opponent). Discipline is part of the key so a 2v1 vs A never
            # merges with a doubles vs A + unnamed.
            opps = sorted(
                [
                    (m.opponent_id, m.opponent.name if m.opponent else "?",
                     _level_of(m.opp_points_snap, m.opponent, my_now)),
                    (m.opponent2_id, m.opponent2.name if m.opponent2 else None,
                     _level_of(m.opp2_points_snap, m.opponent2, my_now)),
                ],
                key=lambda t: (t[1] is None, (t[1] or "").lower()),
            )
            key = f"{m.discipline}|{m.partner_id}|{opps[0][0]}-{opps[1][0]}"
            rec = doubles_h2h.get(key)
            if rec is None:
                rec = doubles_h2h[key] = {
                    **_blank_match_stats(),
                    "key": key,
                    "discipline": m.discipline,
                    "partner_id": m.partner_id,
                    "partner_name": m.partner.name if m.partner else None,
                    "partner_level": _level_of(
                        m.partner_points_snap, m.partner, my_now
                    ),
                    "opp1_id": opps[0][0],
                    "opp1_name": opps[0][1] or "?",
                    "opp1_level": opps[0][2] or "unrated",
                    "opp2_id": opps[1][0],
                    "opp2_name": opps[1][1],
                    "opp2_level": opps[1][2],
                    "last_date": None,
                    "last_result": None,
                }
        else:
            rec = singles_h2h.get(m.opponent_id)
            if rec is None:
                rec = singles_h2h[m.opponent_id] = {
                    **_blank_match_stats(),
                    "opponent_id": m.opponent_id,
                    "name": m.opponent.name if m.opponent else "?",
                    # Derived at-match-time relative level (snapshot points vs
                    # my current rating) — the h2h record's label.
                    "level": _level_of(m.opp_points_snap, m.opponent, my_now)
                    or "unrated",
                    "last_date": None,
                    "last_result": None,
                    "matches": [],
                }
        rec.setdefault("matches", [])
        _tally(rec, m)
        # matches are ordered by date, so the last seen is the most recent.
        rec["last_date"] = m.date
        rec["last_result"] = _result_of(m)
        rec["matches"].append(
            schemas.MatchLine(
                date=m.date,
                discipline=m.discipline,
                my_sets=m.my_sets,
                opp_sets=m.opp_sets,
                result=_result_of(m),
                handicap=m.handicap or 0,
                handicap_pattern=m.handicap_pattern,
                event_name=m.event.name if m.event else None,
            )
        )
    return _H2HAcc(overall, singles_h2h, doubles_h2h, opp_brief)


# Rolling "form": win rate over the last FORM_WINDOW decided (W/L) matches.
# Per-bucket win rate is pure noise at day granularity (2-3 matches/day), so
# the trend chart plots this window instead; FORM_MIN keeps the very first
# points from being a meaningless 0%/100% off one or two matches.
FORM_WINDOW = 10
FORM_MIN = 3


def _prior_form_results(
    db: Session, date_from: dt.date, discipline: str, category: str
) -> list[str]:
    """W/L letters of the last FORM_WINDOW decided named matches before
    `date_from` (never before the tab floor), oldest first — they seed the
    rolling form so the line doesn't restart from scratch at the range edge."""
    if date_from <= MATCH_STATS_FLOOR:
        return []
    earlier = _query_named_matches(
        db,
        MATCH_STATS_FLOOR,
        date_from - dt.timedelta(days=1),
        discipline,
        category,
        with_relations=False,  # only scores are read — skip the eager loads
    )
    results = [r for r in map(_result_of, earlier) if r != "T"]
    return results[-FORM_WINDOW:]


def _trend_buckets(
    matches: list[Match],
    date_from: dt.date,
    date_to: dt.date,
    unit: str,
    prior_results: list[str] | None = None,
) -> list[schemas.MatchTrendBucket]:
    """W/L counts per day/week/month bucket + the rolling form at each end."""
    by_iso: dict[str, list[Match]] = {}
    for m in matches:
        by_iso.setdefault(m.date.isoformat(), []).append(m)
    window: deque[str] = deque(prior_results or (), maxlen=FORM_WINDOW)
    trend: list[schemas.MatchTrendBucket] = []
    for key, label, b_from, b_to in _bucket_ranges(date_from, date_to, unit):
        b_m = b_w = b_l = 0
        for d in _date_range(b_from, b_to):
            for m in by_iso.get(d.isoformat(), []):
                b_m += 1
                r = _result_of(m)
                if r == "W":
                    b_w += 1
                elif r == "L":
                    b_l += 1
                if r != "T":
                    window.append(r)
        trend.append(
            schemas.MatchTrendBucket(
                key=key,
                label=label,
                date_from=b_from,
                date_to=b_to,
                matches=b_m,
                wins=b_w,
                losses=b_l,
                win_rate=win_rate(b_w, b_l),
                form=(
                    window.count("W") / len(window)
                    if len(window) >= FORM_MIN
                    else None
                ),
            )
        )
    return trend


def build_match_stats(
    db: Session,
    date_from: dt.date,
    date_to: dt.date,
    discipline: str = "all",
    category: str = "all",
    unit: str = "month",
    replay: rating.ReplayResult | None = None,
    form_seed: bool = True,
) -> schemas.MatchStatsResponse:
    """Stats over *named-opponent* matches only (opponent_id set, playing).

    Unlike build_stats (which counts every match), this tab is opponent-centric,
    so matches without a recorded opponent are excluded. A match is attributed to
    its primary opponent_id (opponent #1 in doubles); use the Singles filter for
    clean 1-v-1 analysis.

    The trend buckets carry the rolling form; seeding it issues one extra
    pre-range query (`_prior_form_results`) — callers that ignore `form`
    (the coach bundle) pass `form_seed=False` to skip it.
    """
    # Clamp to the floor — this tab only covers matches from June 2026 on.
    date_from = max(date_from, MATCH_STATS_FLOOR)
    matches = _query_named_matches(db, date_from, date_to, discipline, category)

    # Relative levels derive from POINTS vs my current dynamic rating (the
    # hand-picked label was retired 2026-07-27). Per-MATCH grouping honours
    # the at-match-time snapshot; per-PLAYER listings use current points.
    # `replay` lets a caller doing several ELO-dependent builds (the coach
    # bundle) replay once and share the result.
    my_now = round((replay or rating.replay(db))[0])

    acc = _h2h_accumulate(matches, my_now)
    singles_list = sorted(
        (
            schemas.OpponentRecord(
                opponent_id=r["opponent_id"],
                name=r["name"],
                level=r["level"],
                **_record_tail(r),
            )
            for r in acc.singles_h2h.values()
        ),
        key=lambda o: (-o.played, o.name.lower()),
    )
    doubles_list = sorted(
        (
            schemas.DoublesRecord(
                **{
                    k: r[k]
                    for k in (
                        "key", "discipline", "partner_id", "partner_name",
                        "partner_level", "opp1_id", "opp1_name", "opp1_level",
                        "opp2_id", "opp2_name", "opp2_level",
                    )
                },
                **_record_tail(r),
            )
            for r in acc.doubles_h2h.values()
        ),
        key=lambda r: (-r.played, r.opp1_name.lower()),
    )
    opponents = sorted(
        (schemas.OpponentBrief(**b) for b in acc.opp_brief.values()),
        key=lambda o: (-o.played, o.name.lower()),
    )

    return schemas.MatchStatsResponse(
        date_from=date_from,
        date_to=date_to,
        discipline=discipline,
        category=category,
        unit=unit,
        overall=_finalize_match_stats(acc.overall),
        opponents=opponents,
        singles_h2h=singles_list,
        doubles_h2h=doubles_list,
        trend=_trend_buckets(
            matches, date_from, date_to, unit,
            _prior_form_results(db, date_from, discipline, category)
            if form_seed
            else None,
        ),
    )


def build_handicap_split(
    db: Session,
    date_from: dt.date,
    date_to: dt.date,
    replay: rating.ReplayResult | None = None,
) -> dict[str, dict[str, dict]]:
    """Win rates by opponent level × handicap direction, for the Head Coach.

    Handicap is signed (+N = I give N points per set, -N = I receive), so a
    handicapped match plays out differently from an even one: receiving points
    against a stronger opponent pre-balances the match, while giving points to
    a weaker one puts me at a disadvantage. Win rates must therefore not be
    pooled across the directions. Named-opponent playing matches only, clamped
    to the same floor as build_match_stats. Empty cells are omitted.
    """
    date_from = max(date_from, MATCH_STATS_FLOOR)
    matches = (
        db.query(Match)
        .options(selectinload(Match.opponent))
        .filter(
            Match.date >= date_from,
            Match.date <= date_to,
            Match.is_nonplaying == False,  # noqa: E712
            Match.opponent_id.isnot(None),
        )
        .all()
    )
    my_now = round((replay or rating.replay(db))[0])
    acc: dict[str, dict[str, dict]] = {
        lv: {d: _blank_match_stats() for d in ("even", "receive", "give")}
        for lv in _LEVEL_ORDER
    }
    for m in matches:
        lv = _level_of(m.opp_points_snap, m.opponent, my_now) or "unrated"
        if lv not in acc:
            continue
        h = m.handicap or 0
        direction = "even" if h == 0 else ("give" if h > 0 else "receive")
        _tally(acc[lv][direction], m)
    return {
        lv: {
            d: {
                "played": s["total"],
                "wins": s["wins"],
                "losses": s["losses"],
                "win_rate": win_rate(s["wins"], s["losses"]),
            }
            for d, s in dirs.items()
            if s["total"]
        }
        for lv, dirs in acc.items()
    }


# ------------------------------------------------------------ ELO breakdown


def build_rating_breakdown(
    db: Session,
    date_from: dt.date,
    date_to: dt.date,
    unit: str = "day",
    replay: rating.ReplayResult | None = None,
    with_movers: bool = True,
) -> schemas.MyRatingBreakdownOut:
    """Net ELO change per day/week/month bucket + every counted match's ±Δ.

    Replayed on demand via rating.replay — nothing stored. The rating is
    GLOBAL: no discipline/category filtering here (deltas could be filtered,
    but a filtered "rating at end of bucket" would be a lie — decided v1
    2026-07-27). `with_movers=False` skips building the per-match rows for
    callers that only read the buckets (the coach bundle)."""
    _final, steps = replay or rating.replay(db)
    anchor_date = rating.get_my_anchor_date(db)
    anchor_points = rating.get_my_points(db)

    def value_at(day: dt.date) -> int | None:
        """Replayed rating at the END of `day` (carry-forward over quiet
        days); None before the anchor, when no rating existed yet."""
        if day < anchor_date:
            return None
        val = float(anchor_points)
        for s in steps:
            if s.date > day:
                break
            val = s.rating_after
        return round(val)

    buckets: list[schemas.RatingBucketOut] = []
    for key, label, b_from, b_to in _bucket_ranges(date_from, date_to, unit):
        in_bucket = [s for s in steps if b_from <= s.date <= b_to]
        buckets.append(
            schemas.RatingBucketOut(
                key=key,
                label=label,
                date_from=b_from,
                date_to=b_to,
                delta=round(sum(s.delta for s in in_bucket), 1),
                counted=len(in_bucket),
                rating_end=value_at(b_to),
            )
        )

    in_range = [s for s in steps if date_from <= s.date <= date_to]

    # EVERY counted match becomes a table row (the GUI sorts client-side).
    # One date-window query batch-loads them (a superset that includes
    # skipped matches — harmless, and it avoids an unbounded IN(id...) list).
    match_by_id: dict[int, Match] = (
        {
            m.id: m
            for m in _playing_matches(db).filter(
                Match.date >= date_from, Match.date <= date_to
            )
        }
        if with_movers and in_range
        else {}
    )

    def _mover(s: "rating.ReplayStep") -> schemas.RatingMoverOut:
        m = match_by_id.get(s.match_id)
        return schemas.RatingMoverOut(
            match_id=s.match_id,
            date=s.date,
            delta=round(s.delta, 1),
            discipline=m.discipline if m else "singles",
            opponent_name=m.opponent.name if m and m.opponent else None,
            opponent2_name=m.opponent2.name if m and m.opponent2 else None,
            partner_name=m.partner.name if m and m.partner else None,
            my_sets=m.my_sets if m else 0,
            opp_sets=m.opp_sets if m else 0,
        )

    # Newest first — the GUI table's default sort order.
    movers = [_mover(s) for s in reversed(in_range)] if with_movers else []

    # Rating carried INTO the range; the anchor itself counts as the start
    # when the range begins on/before the anchor day.
    rating_start = value_at(date_from - dt.timedelta(days=1))
    if rating_start is None and date_to >= anchor_date:
        rating_start = anchor_points

    return schemas.MyRatingBreakdownOut(
        date_from=date_from,
        date_to=date_to,
        unit=unit,
        anchor_date=anchor_date,
        anchor_points=anchor_points,
        total_delta=round(sum(s.delta for s in in_range), 1),
        counted=len(in_range),
        rating_start=rating_start,
        rating_end=value_at(date_to),
        buckets=buckets,
        movers=movers,
    )


# ---------------------------------------------------------------- export

_RATING_HEX = {"green": "FF63BE7B", "yellow": "FFFFEB84", "red": "FFE06666"}
_COLOR_HEX = {"green": "FF92D050", "yellow": "FFFFFF00", "none": None}


def _date_range(date_from: dt.date, date_to: dt.date) -> list[dt.date]:
    n = (date_to - date_from).days
    return [date_from + dt.timedelta(days=i) for i in range(n + 1)]


def _build_grid(db: Session, date_from: dt.date, date_to: dt.date):
    """Return (categories, dates, cell_text, rating_by_date, cell_colors).

    Same _grid_cells renderer as the on-screen grid, so the export can never
    silently drop information the grid shows."""
    dates = _date_range(date_from, date_to)
    rng = _load_range(db, date_from, date_to)
    text, cell_colors, rating_by_date = _grid_cells(db, rng, dates, for_export=True)
    return rng.categories, dates, text, rating_by_date, cell_colors


def export_csv(db: Session, date_from: dt.date, date_to: dt.date) -> bytes:
    categories, dates, text, rating_by_date, _colors = _build_grid(
        db, date_from, date_to
    )
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Category", *[d.isoformat() for d in dates]])
    for c in categories:
        row = [c.label]
        for d in dates:
            iso = d.isoformat()
            if c.key == "overall":
                row.append(rating_by_date.get(iso, ""))
            else:
                row.append(text.get((c.id, iso), "").replace("\n", " | "))
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(db: Session, date_from: dt.date, date_to: dt.date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    categories, dates, text, rating_by_date, cell_colors = _build_grid(
        db, date_from, date_to
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Tracker"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=1, column=1, value="Category").font = header_font
    for j, d in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=j, value=d.strftime("%a %d %b %Y"))
        cell.font = header_font

    for i, c in enumerate(categories, start=2):
        label_cell = ws.cell(row=i, column=1, value=c.label)
        label_cell.font = header_font
        hex_color = _COLOR_HEX.get(c.color_group)
        if hex_color:
            label_cell.fill = PatternFill("solid", fgColor=hex_color)
        for j, d in enumerate(dates, start=2):
            iso = d.isoformat()
            cell = ws.cell(row=i, column=j)
            if c.key == "overall":
                rating = rating_by_date.get(iso)
                if rating and rating in _RATING_HEX:
                    cell.fill = PatternFill("solid", fgColor=_RATING_HEX[rating])
            else:
                cell.value = text.get((c.id, iso), "")
                cell.alignment = wrap
                fill = cell_colors.get((c.id, iso))
                if fill and fill in _RATING_HEX:
                    cell.fill = PatternFill("solid", fgColor=_RATING_HEX[fill])

    ws.column_dimensions["A"].width = 26
    for col in range(2, len(dates) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
